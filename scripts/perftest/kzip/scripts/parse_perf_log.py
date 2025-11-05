import re
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def parse_filename(filename):
    """从文件名提取blocksize和inflight参数"""
    blocksize_match = re.search(r'blocksize-(\d+)', filename)
    inflight_match = re.search(r'inflight-(\d+)', filename)

    blocksize = blocksize_match.group(1) if blocksize_match else 'N/A'
    inflight = inflight_match.group(1) if inflight_match else 'N/A'

    return int(blocksize), int(inflight)

def extract_log_data(log_text, filename):
    blocksize, inflight = parse_filename(filename)

    pattern = r"console\.log\('kae-threads:(\d+).*?multi:1,file:.*?/([^/]*?\.tar(\.compressed)?).*?inflightNum:\s*(\d+).*?alg:([^ ]+).*?\);t\(\'(.*?)\)"
    matches = re.findall(pattern, log_text)

    data_dict = {}

    for match in matches:
        kae_threads, full_filename, compressed_mark, inflight_num, alg, data_part = match
        file_key = "decompress" if compressed_mark else "compress"

        parts = data_part.split()

        # 解析总用时（移除's'并转换为浮点数）
        time_used = float(parts[7].replace('s', ''))

        # 过滤总用时小于1秒的数据
        if time_used < 1:
            continue

        if file_key not in data_dict:
            data_dict[file_key] = []

        row = [
            blocksize,          # 新增：来自文件名的blocksize
            inflight,          # 新增：来自文件名的inflight
            alg,
            int(kae_threads),
            parts[0] + "kb",
            int(inflight_num),
            float(parts[2]),
            parts[3],
            time_used,        # 已经处理过的总用时
            float(parts[8].replace('us', '')),
            float(parts[9].replace('us', '')),
            float(parts[10].replace('us', '')),
            float(parts[11].replace('us', '')),
            float(parts[12].replace('us', '')),
            float(parts[13].replace('us', '')),
            float(parts[14].replace('us', '')),
            filename
        ]
        data_dict[file_key].append(row)
        if parts[0] == "16" or parts[0] == "64":
            # 现在有15列了（原13列+新增的2列）
            data_dict[file_key].append([None]*16)

    return data_dict

def ensure_xlsx_suffix(filename):
    if not filename.lower().endswith('.xlsx'):
        return filename + '.xlsx'
    return filename

def style_excel(output_file):
    wb = load_workbook(output_file)

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(left=Side(style='thin'),
                  right=Side(style='thin'),
                  top=Side(style='thin'),
                  bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    even_fill = PatternFill("solid", fgColor="DCE6F1")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # 设置列宽（新增两列后调整）
        col_widths = [12, 12, 12, 15, 10, 12, 10, 12, 10, 12, 12, 12, 10, 10, 10]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 应用样式
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = alignment

                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.fill = even_fill if cell.row % 2 == 0 else odd_fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)

def create_excel(data_dict, output_file):
    output_file = ensure_xlsx_suffix(output_file)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for file_type, data in data_dict.items():
            columns = [
                "后台blocksize",    # 新增列
                "后台inflight",    # 新增列
                "算法(alg)",
                "压缩等级",
                "包长大小",
                "inflightNum",
                "压缩率",
                "带宽",
                "总用时(s)",   # 单位改为秒
                "平均时延(μs)",
                "最大时延(μs)",
                "最小时延(μs)",
                "P50(μs)",
                "P90(μs)",
                "P99(μs)",
                "P999(μs)",
                "原始文件"
            ]
            sheet_name = f"{file_type}_数据"[:31]

            df = pd.DataFrame(data, columns=columns)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    style_excel(output_file)
    print(f"成功生成美化Excel文件：{output_file}")

def main():
    if len(sys.argv) < 2:
        print("错误：请指定日志文件名作为参数")
        print("用法：python script.py <日志文件名> [输出文件名]")
        sys.exit(1)

    log_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else sys.argv[1]

    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            log_text = f.read()

        # 提取并处理数据（传入文件名用于提取参数）
        data_dict = extract_log_data(log_text, log_file)

        if not data_dict:
            print("警告：未找到任何有效数据（或所有数据总用时均小于1秒）")
            sys.exit(0)

        create_excel(data_dict, output_file)

    except Exception as e:
        print(f"处理文件错误: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
